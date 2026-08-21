# -*- coding: utf-8 -*-
"""表格导入解析: 识别 名称/biz/链接 三列, 支持别名和无表头特征识别"""
import csv
import io
import re

_BASE_DIR = None

# 列别名(有表头时)
NAME_KEYS = {"公众号名称", "公众号", "名称", "name", "账号", "公众号名", "公众号昵称"}
BIZ_KEYS = {"biz", "biz代码", "biz_code", "biz code", "代码", "公众号id", "bizid", "公众号biz", "公众号ID"}
LINK_KEYS = {"链接", "文章链接", "url", "link", "文章url", "地址", "文章地址", "文章链接url"}

BIZ_RE = re.compile(r"^m[A-Za-z0-9+/=_-]{5,}={1,2}$", re.I)   # m开头, 1/2个=结尾
LINK_RE = re.compile(r"mp\.weixin\.qq\.com|^https?://")


def _is_biz(v):
    v = (v or "").strip()
    return bool(BIZ_RE.match(v))


def _is_link(v):
    v = (v or "").strip()
    return bool(LINK_RE.search(v))


def _is_name(v):
    v = (v or "").strip()
    if not v:
        return False
    return not _is_biz(v) and not _is_link(v) and len(v) >= 2


def parse_file(filename, raw):
    """解析文件内容为行字典列表
    filename: 文件名(判后缀); raw: 字节
    返回 [{name,biz,link}, ...] (含缺项的行, name可能为空)"""
    ext = (filename or "").lower().rsplit(".", 1)[-1]
    if ext in ("xlsx", "xlsm"):
        return _parse_excel(raw)
    # 默认 csv (也尝试 tab 分隔等)
    return _parse_csv(raw)


def _parse_csv(raw):
    text = raw.decode("utf-8-sig", errors="replace")
    lines = text.splitlines()
    if not lines:
        return []
    rows = list(csv.reader(lines))
    return _extract(rows)


def _parse_excel(raw):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    return _extract(rows)


def _extract(rows):
    """从原始行列中识别3列并输出列表"""
    rows = [[("" if c is None else str(c)).strip() for c in r if c is not None] if r else [] for r in rows]
    rows = [r for r in rows if any(x.strip() for x in r)]
    if not rows:
        return []
    # 判断第一行是否表头(有名字段被识别)
    head = rows[0]
    head_map = _match_header(head)
    if head_map:
        body = rows[1:]
    else:
        body = rows
    result = []
    for r in body:
        item = {"name": "", "biz": "", "link": ""}
        if head_map:
            # 有表头: 按列映射
            for i, v in enumerate(r):
                if i >= len(r):
                    break
                role = head_map.get(i)
                if role and v:
                    item[role] = v
        else:
            # 无表头: 逐格特征识别
            for v in r:
                if not v:
                    continue
                if _is_biz(v) and not item["biz"]:
                    item["biz"] = v
                elif _is_link(v) and not item["link"]:
                    item["link"] = v
                elif _is_name(v) and not item["name"]:
                    item["name"] = v
        result.append(item)
    return result


def _match_header(head):
    """有表头: 按别名匹配三列; 全部匹配到才返回"""
    mapping = {}
    found = set()
    norm = [str(h).strip().lower() for h in head]
    for i, h in enumerate(norm):
        if h in {k.lower() for k in NAME_KEYS}:
            mapping[i] = "name"
        elif h in {k.lower() for k in BIZ_KEYS}:
            mapping[i] = "biz"
        elif h in {k.lower() for k in LINK_KEYS}:
            mapping[i] = "link"
    if "name" in mapping.values() and ("biz" in mapping.values() or "link" in mapping.values()):
        return mapping
    return None


def _match_by_feature(rows):
    """无表头: 用列特征识别(biz正则/链接); 名称列=既非biz也非链接的列"
    返回 {列index: role}; 列数按最多行扩展"""
    ncols = max(len(r) for r in rows)
    mapping = {}
    col_scores = {i: [] for i in range(ncols)}
    for r in rows:
        for i, v in enumerate(r[:ncols]):
            if _is_biz(v):
                col_scores[i].append("biz")
            elif _is_link(v):
                col_scores[i].append("link")
            elif _is_name(v):
                col_scores[i].append("name")
    for i in range(ncols):
        cnt = col_scores[i]
        if cnt:
            # 该列多数情况是什么
            from collections import Counter
            role = Counter(cnt).most_common(1)[0][0]
            # 名称列: name 出现最多且不是biz/link
            mapping[i] = role
    return mapping


# ---- 文章导入: 提取文章链接 ----
ART_LINK_RE = re.compile(r"https?://(?:mp\.weixin\.qq\.com|weixin\.qq\.com)/\S+", re.I)


def _extract_links(cell):
    """从一个单元格提取文章链接(正则, 安全过滤)"""
    v = "" if cell is None else str(cell).strip()
    if not v or len(v) > 2048:
        return []
    # 仅接受微信文章域名
    if not re.search(r"mp\.weixin\.qq\.com", v, re.I):
        return []
    found = ART_LINK_RE.findall(v)
    # 清理尾部标点
    out = []
    for f in found:
        f = f.rstrip("),.;，。）")
        if f not in out:
            out.append(f)
    return out


def parse_article_file(filename, raw, max_rows=2000):
    """解析上传表格, 提取所有文章链接(去重, 保序)
    返回 [link, ...]; 安全: 仅识别文档表格类型/限定行数/仅微信域名"""
    ext = (filename or "").lower().rsplit(".", 1)[-1]
    if ext not in ("csv", "xlsx", "xlsm", "xls"):
        raise ValueError("仅支持 CSV / Excel 文件")
    if ext in ("xlsx", "xlsm", "xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    else:
        text = raw.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(text.splitlines()))
    links = []
    seen = set()
    for r in rows:
        for c in r:
            if c is None:
                continue
            for lk in _extract_links(c):
                if lk not in seen:
                    seen.add(lk)
                    links.append(lk)
        if len(links) >= max_rows:
            break
    return links


# ---- 文章多列导入: 识别表头映射各列字段 ----
ART_COLUMN_ALIASES = {
    "title": {"标题", "title", "文章标题", "名称", "标题名"},
    "date": {"日期", "date", "发布时间", "时间", "发布日期"},
    "link": {"链接", "link", "url", "文章链接", "地址", "文章url"},
    "biz": {"biz", "biz代码", "biz_code", "公众号id", "公众号ID", "公众号biz", "bizid"},
    "reads": {"阅读", "阅读量", "reads", "阅读数"},
    "likes": {"点赞", "点赞量", "likes", "赞"},
    "forwards": {"转发", "转发量", "forwards", "转发数"},
    "favorites": {"喜欢", "喜欢量", "favorites", "在看", "在看数"},
    "comments": {"评论", "评论量", "comments", "评论数"},
    "original": {"原创", "是否原创", "原创标识", "original"},
    "ip": {"ip", "IP", "IP属地", "属地"},
}

# 反向: 规范化别名 -> 字段
ALIAS_TO_FIELD = {}
for _field, _aliases in ART_COLUMN_ALIASES.items():
    for _a in _aliases:
        ALIAS_TO_FIELD[_a.lower()] = _field


def _norm_date(v):
    v = (v or "").strip()
    m = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})", v)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return v


def parse_article_rows(filename, raw, max_rows=2000):
    """识别表头多列(标题/日期/链接/阅读/点赞等), 返回行dict列表
    返回 [{"title","date","link","biz","reads","likes","forwards","favorites","comments","original","ip"}, ...]"""
    ext = (filename or "").lower().rsplit(".", 1)[-1]
    if ext not in ("csv", "xlsx", "xlsm", "xls"):
        raise ValueError("仅支持 CSV / Excel 文件")
    if ext in ("xlsx", "xlsm", "xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        grid = [[("" if c is None else str(c)) for c in row] for row in ws.iter_rows(values_only=True)]
    else:
        text = raw.decode("utf-8-sig", errors="replace")
        grid = [list(r) for r in csv.reader(text.splitlines())]
    grid = [r for r in grid if any(x.strip() for x in r)]
    if not grid:
        return []
    # 第一行识别表头
    head = grid[0]
    colmap = {}   # 列index -> 字段
    for i, h in enumerate(head):
        f = ALIAS_TO_FIELD.get(str(h).strip().lower())
        if f:
            colmap[i] = f
    # 无表头时通过链接正则识别链接列
    body = grid[1:] if colmap else grid
    if not colmap:
        for i, r in enumerate(body):
            for j, c in enumerate(r):
                if _is_link(c) and "link" not in colmap.values():
                    colmap[j] = "link"
    # 按列index排序字段, 便于读取
    col_items = sorted(colmap.items())
    rows = []
    for r in body[:max_rows]:
        item = {k: "" for k in ART_COLUMN_ALIASES}
        for i, field in col_items:
            if i < len(r):
                v = r[i].strip()
                if field == "date":
                    v = _norm_date(v)
                if v:
                    item[field] = v
        # 至少有链接才视为有效行
        if item.get("link"):
            rows.append(item)
    return rows
